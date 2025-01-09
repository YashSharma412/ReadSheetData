import json
from openpyxl import load_workbook, Workbook
from pprint import pprint

# Define file paths
input_file_path = "./test/testInputFiles/input.xlsm"
output_file_path = "./test/testOutputFiles/output.xlsm"
mapping_file_path = "mapping.json"

def read_and_display_data(input_path):
    # Load the workbook and select the active sheet
    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active
    
    # Iterate through the rows and print the values
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
        # print(row)
    
    return data

def process_and_write_data(data, mapping_path, output_path):
    # Load the mapping schema
    with open(mapping_path, 'r') as file:
        schema = json.load(file)
    # Iniyialize function vars
    main_table = schema['input_data']['main_table']
    start_cell = main_table['start_cell']
    end_row = main_table['end_row']
    columns = main_table['columns']
    
    # Extract the starting row and column from the start_cell
    start_row = int(start_cell[1:])
    start_col = start_cell[0]
    
    # Detect the start of data dynamically (neglect headers or empty rows at start)
    start_data_row = start_row
    for i, row in enumerate(data[start_row-1:], start=start_row):
        if any(row[ord(col) - ord('A')] is not None for col in columns):
            start_data_row = i
            break
    
    # Create a dictionary to store the data
    table_data = {columns[col]: [] for col in columns}
    
    # Populate the dictionary with data
    for row in data[start_data_row:end_row]:
        # Skip rows that are entirely None
        if all(row[ord(col) - ord('A')] is None for col in columns):
            continue
        for col, col_name in columns.items():
            col_index = ord(col) - ord('A')
            table_data[col_name].append(row[col_index])
    
    # Pretty print the dictionary data
    pprint(table_data);
    
    # Write the data to the output file
    write_output_data(table_data, schema['output_data']['main_table'], output_path)

def write_output_data(table_data, output_schema, output_path):
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    
    # Extract the starting cell and columns from the output schema
    start_cell = output_schema['start_cell']
    start_row = int(start_cell[1:])
    start_col = start_cell[0]
    columns = output_schema['columns']
    
    # Write the headers
    for col, col_name in columns.items():
        col_index = ord(col) - ord('A')
        sheet.cell(row=start_row, column=col_index + 1, value=col_name)
    
    # Write the data
    for i, row_data in enumerate(zip(*table_data.values()), start=start_row + 1):
        for col, value in zip(columns.keys(), row_data):
            col_index = ord(col) - ord('A')
            sheet.cell(row=i, column=col_index + 1, value=value)
    
    # Save the workbook to the output path
    workbook.save(output_path)

# Call the function to read and display data
data = read_and_display_data(input_file_path)

# Call the function to process and write data
process_and_write_data(data, mapping_file_path, output_file_path)