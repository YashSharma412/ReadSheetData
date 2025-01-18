import json
from openpyxl import load_workbook, Workbook
import os
from pathlib import Path
import shutil
from copy import copy  # Import the copy function

# Configuration that will later come from database
# List of input files to be merged
INPUT_FILES = [
    {"path": "./roaster/input/input1.xlsx", "type": "excel"},
    {"path": "./roaster/input/input2.xlsx", "type": "excel"},
]

# Path to mapping schema that contains merge configuration
mapping_file_path = "./mappings/roaster-mapping.json"

# Load the merge configuration globally
with open(mapping_file_path, "r") as file:
    mapping_schema = json.load(file)
merge_config = mapping_schema.get("merge_files", [{}])[0]

def get_sheet(workbook, sheet_name):
    """
    Get the specified sheet from the workbook.
    If sheet_name is 'active', return the active sheet.
    If sheet_name is not found, log a warning and return the active sheet.
    """
    if sheet_name == "active":
        return workbook.active
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    print(f"Warning: Sheet '{sheet_name}' not found. Using active sheet as default.")
    return workbook.active

def validate_template_format(workbooks):
    """
    Compare the headers and structure of workbooks to ensure they match.
        Args: 
            workbooks: List of loaded Excel workbooks
        Returns:
            bool: True if all workbooks have matching headers, False otherwise
    """
    if not workbooks:
        return False

    print("Validating template format...")
    # Get which row contains headers (default is row 1) from json config file
    header_row = merge_config.get("header_row", 1)

    # Use first workbook's headers as reference for comparison
    reference_wb = workbooks[0]
    reference_sheet = get_sheet(reference_wb, merge_config.get("sheet", "active"))
    reference_headers = [cell.value for cell in reference_sheet[header_row]]
    print(f"Reference headers: {reference_headers}")

    # Compare each workbook's headers with reference
    for wb in workbooks[1:]:
        sheet = get_sheet(wb, merge_config.get("sheet", "active"))
        current_headers = [cell.value for cell in sheet[header_row]]
        print(f"Current headers: {current_headers}")
        if current_headers != reference_headers:
            print("Headers do not match across files. The templates must be of different types.")
            return False

    print("All headers match.");
    return True

def copy_sheet(source_sheet, target_sheet):
    """
    Copy all data, formatting, and formulas from source_sheet to target_sheet.
    """
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

def merge_workbooks(workbooks, template_path, output_file_path):
    """
    Merge multiple workbooks into a single workbook.Preserves headers from first workbook and concatenates all data rows.

    Args:
        workbooks: List of workbooks to merge
        template_path: Path to the template file
        output_file_path: Path to save the merged workbook

    Returns:
        Workbook: New workbook containing merged data, or None if error
    """
    if not workbooks:
        return None

    print("Merging workbooks...")

    # Create new workbook for merged data by copying the reference workbook
    shutil.copy2(template_path, output_file_path)
    merged_wb = load_workbook(output_file_path)
    merged_sheet = merged_wb.active

    # Get row numbers for headers and data start
    header_row = merge_config.get("header_row", 1)
    data_start_row = merge_config.get("data_start_row", 2)

    # Track where to insert next row in merged sheet
    current_row = merged_sheet.max_row + 1

    # Calculate the maximum number of columns filled with data
    max_cols = merged_sheet.max_column
    print(f"Max columns: {max_cols}")

    # Merge data from all workbooks except the reference workbook
    for wb in workbooks[1:]:
        sheet = get_sheet(wb, merge_config.get("sheet", "active"))
        max_row = sheet.max_row
        print(f"Processing sheet: {sheet.title}, Max rows: {max_row}")

        # Copy all non-empty data rows from current workbook
        for row_idx in range(data_start_row, max_row + 1):
            row = sheet[row_idx]
            if all(cell.value is None for cell in row):
                continue  # Skip completely empty rows
            for col_idx in range(1, max_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                new_cell = merged_sheet.cell(row=current_row, column=col_idx, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
            current_row += 1

    print("Merging completed.")
    return merged_wb

def get_next_available_filename(output_dir, base_name="merged_input", ext=".xlsx"):
    """
    Generate unique filename for merged output.
    Avoids overwriting existing files by adding numeric suffix.

    Args:
        output_dir: Directory to save file
        base_name: Base filename without extension
        ext: File extension

    Returns:
        str: Full path to next available filename
    """
    i = 0
    while True:
        suffix = str(i) if i > 0 else ""
        filename = f"{base_name}{suffix}{ext}"
        filepath = os.path.join(output_dir, filename)
        if not os.path.exists(filepath):
            return filepath
        i += 1

def process_input_files():
    """
    Main function that:
        1. Loads all input workbooks
        2. Validates they have same template format
        3. Merges them into single workbook
        4. Saves merged workbook to output directory
    Returns:
        str: Path to merged file if successful, None if error
    """
    print("Processing input files...")
    
    # Create output directory if it doesn't exist
    output_dir = "./roaster/input/merged"
    os.makedirs(output_dir, exist_ok=True)

    # Load all input workbooks
    workbooks = []
    for input_file in INPUT_FILES:
        try:
            print(f"Loading workbook: {input_file['path']}")
            wb = load_workbook(input_file["path"], data_only=True)
            workbooks.append(wb)
        except Exception as e:
            print(f"Error loading {input_file['path']}: {e}")
            return None

    # First validate all workbooks have same format
    if not validate_template_format(workbooks):
        print("Error: Input files have different formats")
        return None

    # If validation passed, merge the workbooks
    template_path = INPUT_FILES[0]["path"]
    output_path = get_next_available_filename(output_dir)
    merged_wb = merge_workbooks(workbooks, template_path, output_path)
    if not merged_wb:
        print("Error: Failed to merge workbooks")
        return None

    # Save merged workbook
    merged_wb.save(output_path)
    print(f"Successfully merged input files to: {output_path}")
    return output_path

if __name__ == "__main__":
    # Run merge process
    merged_file_path = process_input_files()
    if merged_file_path:
        print("Files merged successfully")
    else:
        print("Failed to merge files")
