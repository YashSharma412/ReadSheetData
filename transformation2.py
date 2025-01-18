import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
from pprint import pprint
import sys
from difflib import get_close_matches
import os
import shutil
from pathlib import Path

# Configuration that will later come from database
INPUT_FILES = [
    {"path": "./roaster/input/input1.xlsx", "type": "excel"},
    {"path": "./roaster/input/input2.xlsx", "type": "excel"}
]
TEMPLATE_FILE = {"path": "./ref/output_template.xlsx", "type": "excel"}
OUTPUT_PATH = "./roaster/output/"

# File paths for mappings
mapping_file_path = "./mappings/roaster-mapping.json"
data_mapping_file_path = "./mappings/data-mappings.json"

# Load the mapping schemas globally
with open(mapping_file_path, "r") as file:
    mapping_schema = json.load(file)

with open(data_mapping_file_path, "r") as file:
    data_mappings = json.load(file)

# Define transformation functions
def split_name(name, separator=" "):
    if isinstance(name, list):
        return [split_name(n, separator) for n in name]
    
    parts = [part for part in name.strip().split(separator) if part]
    if not parts:
        return ["", ""]
    if len(parts) == 1:
        return [parts[0], ""]
    return [parts[0], parts[-1]]

def capitalize(text):
    if isinstance(text, list):
        return [capitalize(sub_text) for sub_text in text]
    return text.capitalize()

def convert_to_integer(text):
    if isinstance(text, list):
        return [convert_to_integer(t) for t in text]
    try:
        return int(text)
    except ValueError:
        return None

def uppercase(text):
    if isinstance(text, list):
        return [uppercase(sub_text) for sub_text in text]
    return text.upper()

def title_case(text):
    """Convert text to title case, handling edge cases and arrays"""
    # Handle array/list input
    if isinstance(text, list):
        if not text:
            return []
        # If input is array, output should also be array
        if isinstance(text[0], list):
            return [title_case(sub_text) for sub_text in text]
        return [title_case(item) for item in text]

    if not isinstance(text, str):
        return text

    # List of words that should remain lowercase
    small_words = {
        "a",
        "an",
        "and",
        "as",
        "at",
        "but",
        "by",
        "for",
        "in",
        "of",
        "on",
        "or",
        "the",
        "to",
        "via",
        "with",
    }

    words = str(text).lower().split()
    if not words:
        return text

    result = []
    for i, word in enumerate(words):
        # Capitalize if it's first word, last word, or not in small_words
        if i == 0 or i == len(words) - 1 or word not in small_words:
            result.append(word.capitalize())
        else:
            result.append(word)

    return " ".join(result)

def data_mapper(text, mapping_key):
    """Replace text with its alternate term from value_mappings using fuzzy matching"""
    if not text:
        return text

    if isinstance(text, list):
        return [data_mapper(item, mapping_key) for item in text]

    # Convert input to string for comparison
    text = str(text)

    # Get the mapping dictionary for the specified key from data_mappings
    value_mappings = data_mappings.get("value_mappings", {}).get(mapping_key, {})

    if not value_mappings:
        return text

    # Convert all mapping keys to strings for comparison
    string_mappings = {str(k): v for k, v in value_mappings.items()}

    # Try exact match first
    if text in string_mappings:
        return string_mappings[text]

    # If no exact match, try fuzzy matching
    try:
        matches = get_close_matches(text, string_mappings.keys(), n=1, cutoff=0.75)
        # If we found a fuzzy match, use its mapping
        if matches:
            return string_mappings[matches[0]]
    except TypeError as e:
        print(f"Warning: Fuzzy matching failed for '{text}' - {str(e)}")
    # If no match found or error occurred, return original text
    return text


def generate_data_based_on(source_data, field_name):
    """
    Generate crew numbers based on classification data.
    source_data: Array of classification values
    field_name: The field whose transformation rules to use
    """
    if not source_data or not isinstance(source_data, list):
        print(f"DEBUG: Empty or invalid source data for {field_name}")
        return []

    rules = mapping_schema.get("transformation_rules", {}).get(field_name, {})
    if not rules or "counter_rules" not in rules:
        print(f"DEBUG: No counter rules found for {field_name}")
        return source_data

    # Get counter configuration
    counter_rules = rules["counter_rules"]
    count = counter_rules.get("start", 1)
    increment_triggers = counter_rules.get("increment_on", [])
    case_sensitive = counter_rules.get("case_sensitive", False)
    output_type = counter_rules.get("output_type", "string")
    # print(f"\nProcessing Crew Numbers:")
    # print(f"Initial count: {count}")
    # print(f"Increment triggers: {increment_triggers}")

    result = []
    previous_value = None
    for value in source_data:
        check_value = str(value) if case_sensitive else str(value).lower()
        check_triggers = (
            increment_triggers
            if case_sensitive
            else [trigger.lower() for trigger in increment_triggers]
        )

        if check_value in check_triggers:
            trigger_index = check_triggers.index(check_value)
            if trigger_index == 0:
                count += 1
                # print(f"Incrementing count to {count} for value: {value}")
            elif trigger_index == 1 and previous_value != check_triggers[0]:
                count += 1
                # print(f"Incrementing count to {count} for value: {value}")

        result.append(str(count) if output_type == "string" else count)
        previous_value = check_value
        # print(f"Value: {value} -> Crew Number: {count}")
    return result

# Map transformation names to functions
transformation_functions = {
    "split_name": split_name,
    "capitalize": capitalize,
    "convert_to_integer": convert_to_integer,
    "uppercase": uppercase,
    "title_case": title_case,
    "data_mapper": data_mapper,
    "generate_data_based_on": generate_data_based_on,
}

def apply_transformations(data, transformations):
    for transformation in transformations:
        func_name = transformation
        params = []

        # Check if transformation has parameters
        if "(" in transformation:
            func_name = transformation.split("(")[0]
            params_str = transformation.split("(")[1].rstrip(")")
            params = [param.strip() for param in params_str.split(",") if param.strip()]

        func = transformation_functions.get(func_name)
        if func:
            print(f"Applying transformation: {func_name} with params: {params}")
            if isinstance(data, list):
                data = func(data, *params)
            else:
                data = func(data, *params)
    return data

def validate_data(data, validations):
    for validation in validations:
        if validation["type"] == "required":
            if isinstance(data, list):
                if not data:  # Empty list is invalid for required field
                    raise ValueError(validation["message"])
                for item in data:
                    if isinstance(item, list):
                        if all(
                            not sub_item
                            or (isinstance(sub_item, str) and not sub_item.strip())
                            for sub_item in item
                        ):
                            raise ValueError(validation["message"])
                    elif not item or (isinstance(item, str) and not item.strip()):
                        raise ValueError(validation["message"])
            elif not data or (isinstance(data, str) and not data.strip()):
                raise ValueError(validation["message"])
        elif validation["type"] == "allow-empty":
            pass

def read_and_validate_data(input_path, mapping_schema):
    workbook = load_workbook(input_path, data_only=True)
    data_store = {}

    # First pass: get the maximum data length from fields with source
    max_data_length = 0
    for mapping in mapping_schema["mappings"]:
        if "source" in mapping:
            source = mapping["source"]
            sheet_name = source["sheet"]
            cell_range = source["range"]
            sheet = workbook[sheet_name]

            if cell_range.endswith("_"):
                start_cell = cell_range.split(":")[0]
                start_row = int(start_cell[1:])
                max_data_length = max(max_data_length, sheet.max_row - start_row + 1)

    # Second pass: process all fields
    for mapping in mapping_schema["mappings"]:
        field_name = mapping["field_name"]
        default_value = mapping.get("default")

        # Special handling for fields that depend on other fields
        if "reference_field" in mapping:
            # Use the referenced field's data for transformations
            reference_field = mapping["reference_field"]
            if reference_field not in data_store:
                raise ValueError(
                    f"Reference field {reference_field} must be processed before {field_name}"
                )
            data = data_store[reference_field]  # Use reference field's data
        else:
            # Normal field processing with source
            if "source" not in mapping:
                if default_value is not None:
                    if (
                        isinstance(mapping["destination"][0]["range"], str)
                        and ":" in mapping["destination"][0]["range"]
                    ):
                        # For range destinations, create list of default values matching max length
                        data = [default_value] * max_data_length
                    else:
                        # For single cell destinations, use single value
                        data = default_value
                    data_store[field_name] = data
                continue

            source = mapping["source"]
            sheet_name = source["sheet"]
            cell_range = source["range"]
            sheet = workbook[sheet_name]

            # Check validation requirements
            is_required = any(
                validation.get("type") == "required"
                for validation in mapping.get("validation", [])
            )
            allows_none = any(
                validation.get("type") == "allow-empty"
                for validation in mapping.get("validation", [])
            )

            # Handle special cases for dynamic ranges
            if cell_range.endswith("_"):
                start_cell = cell_range.split(":")[0]
                col_letter = start_cell[0]
                start_row = int(start_cell[1:])
                data = []
                col_idx = column_index_from_string(col_letter)
                for row in sheet.iter_rows(
                    min_row=start_row,
                    min_col=col_idx,
                    max_col=col_idx,
                    values_only=True,
                ):
                    value = row[0]
                    # Priority order for handling empty/null values:
                    # 1. If value exists, use it
                    # 2. If empty and has default value, use default
                    # 3. If empty, no default, but required - raise error
                    # 4. If empty, no default, allows_none - use empty space
                    # 5. Otherwise keep as null
                    if value is None or (isinstance(value, str) and not value.strip()):
                        if default_value is not None:
                            # Default value takes highest priority for empty fields
                            data.append(default_value)
                        elif is_required:
                            # Required fields must have a value
                            raise ValueError(
                                f"{field_name} is required but found empty value"
                            )
                        elif allows_none:
                            # If field allows empty and has no default, use space
                            data.append(" ")
                        else:
                            # Keep as null if no other conditions apply
                            data.append(value)
                    else:
                        data.append(value)
            else:
                # Handle fixed ranges and single cells
                if ":" in cell_range:
                    start_cell, end_cell = cell_range.split(":")
                    start_col = column_index_from_string(start_cell[0])
                    start_row = int(start_cell[1:])
                    end_col = column_index_from_string(end_cell[0])
                    end_row = int(end_cell[1:])
                    data = []
                    for row in sheet.iter_rows(
                        min_row=start_row,
                        max_row=end_row,
                        min_col=start_col,
                        max_col=end_col,
                        values_only=True,
                    ):
                        if all(cell is None for cell in row):
                            continue
                        data.extend([cell for cell in row if cell is not None])
                else:
                    # Single cell case
                    cell = sheet[cell_range]
                    data = cell.value

        # Validate data if any validations are specified
        if "validation" in mapping:
            # print(f"Validating {field_name}")
            validate_data(data, mapping["validation"])

        # Store the data
        data_store[field_name] = data
    # pprint(data_store)
    return data_store

def apply_transformations_to_data_store(data_store, mapping_schema):
    for field_name, data in data_store.items():
        for mapping in mapping_schema["mappings"]:
            if mapping["field_name"] == field_name:
                if "transformations" in mapping:
                    data_store[field_name] = apply_transformations(
                        data, mapping["transformations"]
                    )
    return data_store


def apply_cell_format(cell, format_config):
    if not format_config:
        return

    # Create Font object
    if "font" in format_config:
        font_config = format_config["font"].copy()
        cell.font = Font(**font_config)

    # Create Fill object
    if "fill" in format_config:
        fill_config = format_config["fill"].copy()
        if "type" in fill_config:
            fill_config["patternType"] = fill_config.pop("type")
        if "color" in fill_config:
            fill_config["fgColor"] = fill_config.pop("color")
        cell.fill = PatternFill(**fill_config)

    # Create Border object
    if "border" in format_config:
        border_config = format_config["border"]
        border_style = border_config.get("style", "thin")
        border_color = border_config.get("color", "000000")
        
        # Only apply borders if style is not "none"
        if border_style.lower() != "none":
            side = Side(style=border_style, color=border_color)
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    # Create Alignment object
    if "alignment" in format_config:
        alignment_config = format_config["alignment"].copy()
        cell.alignment = Alignment(**alignment_config)
        

def format_range(sheet, cell_range, format_config):
    """Apply formatting to a range of cells"""
    if not format_config:
        return
        
    if ":" in cell_range:
        if cell_range.endswith("_"):
            # Handle dynamic ranges
            start_cell = cell_range.split(":")[0]
            col_letter = start_cell[0]
            start_row = int(start_cell[1:])
            
            # Find the last row with data in this column
            last_row = start_row
            col_idx = column_index_from_string(col_letter)
            for row in sheet.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
                if row[0].value is not None:
                    last_row = row[0].row
            
            # Apply formatting from start_row to last_row
            for row in range(start_row, last_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                apply_cell_format(cell, format_config)
        else:
            # Handle fixed ranges
            start_cell, end_cell = cell_range.split(":")
            for row in sheet[cell_range]:
                for cell in row:
                    apply_cell_format(cell, format_config)
    else:
        # Handle single cell
        apply_cell_format(sheet[cell_range], format_config)


def apply_conditional_format(sheet, row_index, format_config, data_store, condition):
    """Apply formatting based on conditions from other fields"""
    if not condition or "when" not in condition:
        return

    field_name = condition["when"]["field"]
    expected_value = condition["when"]["equals"]

    # Get the value from data_store for comparison
    field_data = data_store.get(field_name, [])
    if not field_data or row_index >= len(field_data):
        return

    # Check if condition is met
    if field_data[row_index] == expected_value:
        cell_format = condition["apply"]
        # Convert format keys if needed
        if "fill" in cell_format:
            fill_config = cell_format["fill"].copy()
            if "type" in fill_config:
                fill_config["patternType"] = fill_config.pop("type")
            if "color" in fill_config:
                fill_config["fgColor"] = fill_config.pop("color")
            cell_format["fill"] = fill_config
        return cell_format
    return None


def use_mapping_generate_output(data_store, mapping_schema, output_file_path):
    """Generate an Excel output file based on mapping schema and data store"""
    try:
        output_workbook = load_workbook(output_file_path)
    except FileNotFoundError:
        output_workbook = Workbook()

    # Get default formats
    default_formats = mapping_schema.get("default_formats", {})

    for mapping in mapping_schema["mappings"]:
        field_name = mapping["field_name"]
        destinations = mapping["destination"]
        data = data_store.get(field_name, [])

        for destination in destinations:
            sheet_name = destination["sheet"]
            
            # Get or create sheet
            if sheet_name in output_workbook.sheetnames:
                sheet = output_workbook[sheet_name]
            else:
                sheet = output_workbook.create_sheet(sheet_name)

            # Process each range in the destination
            for idx, cell_range in enumerate(destination["range"].split(",")):
                cell_range = cell_range.strip()  # Remove whitespace
                
                if ":" in cell_range:
                    if cell_range.endswith("_"):
                        # Handle dynamic ranges
                        start_cell = cell_range.split(":")[0]
                        col_letter = start_cell[0]
                        start_row = int(start_cell[1:])
                        
                        # Write data and apply formatting
                        for row_offset, value in enumerate(data):
                            cell = sheet[f"{col_letter}{start_row + row_offset}"]
                            
                            # Handle 2D array data
                            if isinstance(value, list):
                                if idx < len(value):
                                    cell.value = value[idx]
                            else:
                                cell.value = value
                            
                            # Apply base formatting
                            if "format" in destination:
                                apply_cell_format(cell, destination["format"])
                            elif "data_cells" in default_formats:
                                apply_cell_format(cell, default_formats["data_cells"])
                            
                            # Apply conditional formatting if specified
                            if "conditional_format" in destination:
                                cond_format = apply_conditional_format(
                                    sheet,
                                    row_offset,
                                    destination["format"],
                                    data_store,
                                    destination["conditional_format"]
                                )
                                if cond_format:
                                    apply_cell_format(cell, cond_format)
                    
                    else:
                        # Handle fixed ranges
                        start_cell, end_cell = cell_range.split(":")
                        start_col = column_index_from_string(start_cell[0])
                        start_row = int(start_cell[1:])
                        end_row = int(end_cell[1:])
                        
                        # Write data and apply formatting
                        for row_offset, value in enumerate(data):
                            if start_row + row_offset <= end_row:
                                cell = sheet.cell(row=start_row + row_offset, column=start_col)
                                
                                # Handle 2D array data
                                if isinstance(value, list):
                                    if idx < len(value):
                                        cell.value = value[idx]
                                else:
                                    cell.value = value
                                
                                # Apply formatting
                                if "format" in destination:
                                    apply_cell_format(cell, destination["format"])
                                elif "data_cells" in default_formats:
                                    apply_cell_format(cell, default_formats["data_cells"])
                
                else:
                    # Handle single cell
                    cell = sheet[cell_range]
                    if isinstance(data, list) and data:
                        if isinstance(data[0], list) and idx < len(data[0]):
                            cell.value = data[0][idx]
                        else:
                            cell.value = data[0]
                    else:
                        cell.value = data
                    
                    # Apply formatting
                    if "format" in destination:
                        apply_cell_format(cell, destination["format"])
                    elif "data_cells" in default_formats:
                        apply_cell_format(cell, default_formats["data_cells"])

            #* Handle cell merging if specified
            if "merge" in destination:
                merge_range = destination["merge"]
                try:
                    # Unmerge first if already merged
                    if merge_range in sheet.merged_cells:
                        sheet.unmerge_cells(merge_range)
                    
                    sheet.merge_cells(merge_range)
                    # Apply format to merged range
                    if "format" in destination:
                        format_range(sheet, merge_range, destination["format"])
                    elif "data_cells" in default_formats:
                        format_range(sheet, merge_range, default_formats["data_cells"])
                except ValueError as e:
                    print(f"Warning: Could not merge cells {merge_range}: {str(e)}")

    # Save the workbook
    output_workbook.save(output_file_path)

def merge_data_stores(data_stores):
    """Merge multiple data stores into one"""
    merged_store = {}
    for store in data_stores:
        for field_name, data in store.items():
            if field_name not in merged_store:
                merged_store[field_name] = []
            if isinstance(data, list):
                merged_store[field_name].extend(data)
            else:
                merged_store[field_name].append(data)
    return merged_store

def get_next_available_filename(output_dir, base_name="output", ext=".xlsx"):
    """Find next available filename in sequence (output.xlsx, output1.xlsx, etc.)"""
    i = 0
    while True:
        suffix = str(i) if i > 0 else ""
        filename = f"{base_name}{suffix}{ext}"
        filepath = os.path.join(output_dir, filename)
        if not os.path.exists(filepath):
            return filepath
        i += 1

def process_files(mapping_schema):
    # Use configuration variables instead of reading from mapping_schema
    input_files = INPUT_FILES
    template_path = TEMPLATE_FILE.get("path")
    output_dir = OUTPUT_PATH

    # Rest of the function remains the same
    os.makedirs(output_dir, exist_ok=True)
    output_file_path = get_next_available_filename(output_dir)

    try:
        data_stores = []
        for input_file in input_files:
            input_file_path = input_file["path"]
            workbook = load_workbook(input_file_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                max_row = worksheet.max_row
                # print(f"Max rows in {input_file_path} - {sheet_name}: {max_row}")
            data_store = read_and_validate_data(input_file_path, mapping_schema)
            data_stores.append(data_store)           
        # Merge all data stores
        merged_data_store = merge_data_stores(data_stores)

        # Apply transformations to the merged data store
        transformed_data_store = apply_transformations_to_data_store(
            merged_data_store, mapping_schema
        )

        # Copy template file to output location or create a new blank workbook
        if template_path and os.path.exists(template_path):
            shutil.copy2(template_path, output_file_path)
        else:
            Workbook().save(output_file_path)

        # Generate output with transformed data
        use_mapping_generate_output(
            transformed_data_store, mapping_schema, output_file_path
        )
    except Exception as e:
        print(f"Error processing files: {e}")
        if os.path.exists(output_file_path):
            os.remove(output_file_path)
        raise


if __name__ == "__main__":
    # Load the mapping schema
    with open(mapping_file_path, "r") as file:
        mapping_schema = json.load(file)
    # Process all files
    process_files(mapping_schema)