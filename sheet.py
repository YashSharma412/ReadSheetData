import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


blackColor = "000000"
whiteColor = "FFFFFF"
blueColor = "4472C4"
grayColor = "D9D9D9"
greenColor = "C2F0C8"
yellowColor = "FEFF03"
orangeColor = "FFC001"
darkBlue = "166082"
lightBlue = "C1E4F5"
blueBro = "82CAEB"

# default_font = Font(name='Aptos Narrow')
border = Border(left=Side(style='thin', color=blackColor),
                right=Side(style='thin', color=blackColor),
                top=Side(style='thin', color=blackColor),
                bottom=Side(style='thin', color=blackColor))

def merge_and_format_cells(sheet, start_cell, end_cell, value, fill_color=whiteColor, text_color=blackColor, alignment='center', height=None, width=None, font_size=11, family="body", isBorder=False, bold=False):
    sheet.merge_cells(f'{start_cell}:{end_cell}')
    cell = sheet[start_cell]
    cell.value = value
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if family == "head":
        cell.font = Font(color=text_color, size=font_size, name="Aptos Narrow", bold=bold)
    elif family == "line":
        cell.font = Font(color=text_color, size=font_size, name="Cambria", bold=bold)
    elif family == "cambria":
        cell.font = Font(color=text_color, size=font_size, name="Cambria", bold=bold)
    else:
        cell.font = Font(color=text_color, size=font_size, name="Arial", bold=bold)

    cell.alignment = Alignment(horizontal=alignment, vertical='center')
    
    if height:
        sheet.row_dimensions[cell.row].height = height
    if width:
        col_letter = get_column_letter(cell.column)
        sheet.column_dimensions[col_letter].width = width

    if isBorder:
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(f'{start_cell}:{end_cell}')
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).border = border
    
def subTitle(dest_sheet):
    merge_and_format_cells(dest_sheet, 'A4', 'B4', 'Company name', greenColor, height=24, alignment="left",family="head", bold=True, isBorder=True)
    merge_and_format_cells(dest_sheet, 'C4', 'M4', 'Premium Utility Contractor', greenColor, family="head",  isBorder=True, bold=True)
    
def line(dest_sheet):
    merge_and_format_cells(dest_sheet, 'A5', 'B5', 'Storm Date', blueBro, alignment="left",family="head", bold=True)
    # merge_and_format_cells(dest_sheet, 'B5', 'B5', '', lightBlue)
    merge_and_format_cells(dest_sheet, 'C5', 'M5', '', blackColor)

def primaryContact(dest_sheet):
    merge_and_format_cells(dest_sheet, 'A6', 'B7', 'Primary Contact', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C6', 'E6', 'Name', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'F6', 'F6', 'Title', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'G6', 'I6', 'Contact Number', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'J6', 'M6', 'Email', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'A8', 'B8', 'Secondary Contact', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'A9', 'B9', 'Additional Contact', grayColor, alignment="left", isBorder=True, bold=True)

def tableHeading(dest_sheet):
    merge_and_format_cells(dest_sheet, 'A22', 'O22', 'PRE STORM Information Required', yellowColor, family="head", font_size=14, isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'P22', 'W22', 'POST STORM - REQUIRED for Invoice Submission', orangeColor, family="head", font_size=14, isBorder=True, bold=True)

    pre_storm_cols = [
        "Enter \nManually ( Req'd )", "Enter \nManually ( Req'd )", "Enter \nManually ( Req'd )", 
        "Select \nDROP DOWN ( Req'd ) ", "Select \nDROP DOWN", "CoC Select \nDROP DOWN", 
        "Type 10 Digit #\nex. 5558675309", "Select \nDROP DOWN ( Req'd )", "Select \nDROP DOWN\n ( Req'd )", 
        "Enter \nManually", "Enter \nManually", "Select \nDROP DOWN\n ( Req'd )", 
        "Contractor or \nCRC Select \nDROP Down", "Contractor or \nCRC Select \nDROP Down", 
        "Contractor or \nCRC Select \nDROP Down"
    ]

    post_storm_cols = [
        "Contractor or \nCRC Select \nDROP Down", "Contractor or \nCRC Select \nDROP Down", 
        "Contractor or \nCRC Select \nDROP Down", "Contractor or \nCRC Select \nDROP Down", 
        "Contractor or \nCRC Select \nDROP Down", "Contractor or \nCRC Select \nDROP Down", 
        "Contractor or \nCRC Select \nDROP Down", "Contractor or \nCRC Select \nDROP Down"
    ]

    additional_cols = [
        "Company", "Last Name", "First Name", "Crew Number", "Crew Leader/ Restore User", 
        "Permit & Tag?", "Phone Number", "Job classification", "Union/Non-Union", 
        "Equipment Type", "Truck Number", "Gender", "Home Area/State", "Shift Start Time", 
        "Working Area", "PPL Permit Holder Name (if assigned)", "Day 1 Hour Total", 
        "Day 2 Hour Total", "Day 3 Hour Total", "Day 4 Hour Total", "Day 5 Hour Total", 
        "Day 6 Hour Total", "Day 7 Hour Total"
    ]

    start_col = 1
    for i, col in enumerate(pre_storm_cols):
        col_letter = get_column_letter(start_col + i)
        merge_and_format_cells(dest_sheet, f'{col_letter}23', f'{col_letter}23', col, yellowColor, width=25, family="head", isBorder=True, bold=True)

    start_col = 16
    for i, col in enumerate(post_storm_cols):
        col_letter = get_column_letter(start_col + i)
        merge_and_format_cells(dest_sheet, f'{col_letter}23', f'{col_letter}23', col, orangeColor, width=25, family="head", isBorder=True, bold=True)

    start_col = 1
    for i, col in enumerate(additional_cols):
        col_letter = get_column_letter(start_col + i)
        merge_and_format_cells(dest_sheet, f'{col_letter}24', f'{col_letter}24', col, darkBlue, width=25, height=30, family="head", font_size=14, isBorder=True, bold=True)

def otherData(dest_sheet):
    merge_and_format_cells(dest_sheet, 'A11', 'B11', 'Union Status', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C11', 'D11', '', alignment="left", isBorder=True)
    merge_and_format_cells(dest_sheet, 'A12', 'B12', 'Home State', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C12', 'D12', 'CT/PA', alignment="left", isBorder=True)
    merge_and_format_cells(dest_sheet, 'A13', 'B13', 'Traveling from: (City / State)', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C13', 'D13', 'Monroe CT / WILKS-BARRE PA', alignment="left", isBorder=True)

    merge_and_format_cells(dest_sheet, 'A15', 'B15', 'Host Utility: (official use only)', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C15', 'D15', '', alignment="left", isBorder=True)
    merge_and_format_cells(dest_sheet, 'A16', 'B16', 'RMAG: (official use only)', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C16', 'D16', '', alignment="left", isBorder=True)
    merge_and_format_cells(dest_sheet, 'A17', 'B17', 'Naming Convention (official use only)', grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'C17', 'D17', '', alignment="left", isBorder=True)

    merge_and_format_cells(dest_sheet, 'F11', 'I11', 'Total Employee Count',grayColor, alignment="left", isBorder=True, bold=True)
    merge_and_format_cells(dest_sheet, 'F12', 'H12', 'Total - Line (FTE\'s)', alignment="left", isBorder=True, font_size=10, family="cambria")
    merge_and_format_cells(dest_sheet, 'F13', 'H13', 'Total - Line (crews)', alignment="left", isBorder=True, font_size=10, family="cambria")
    merge_and_format_cells(dest_sheet, 'F14', 'H14', 'Total - Electricians', alignment="left", isBorder=True, font_size=10, family="cambria")
    merge_and_format_cells(dest_sheet, 'F15', 'H15', 'Total - Damage Assessors', alignment="left", isBorder=True, font_size=10, family="cambria")
    merge_and_format_cells(dest_sheet, 'F16', 'H16', 'Total-Veg', alignment="left", isBorder=True, font_size=10, family="cambria")
    merge_and_format_cells(dest_sheet, 'F17', 'H17', 'Total Equipment Op', alignment="left", isBorder=True, font_size=10, family="cambria")
    merge_and_format_cells(dest_sheet, 'F18', 'H18', 'Other', alignment="left", isBorder=True, font_size=10, family="cambria")


    merge_and_format_cells(dest_sheet, 'A19', 'W21', '')
    # merge_and_format_cells(dest_sheet, 'E10', 'E21', '')
    
    
    merge_and_format_cells(dest_sheet, 'K11', 'M11', 'Total Equipment',grayColor, alignment="left", isBorder=True, bold=True, font_size=10)

    merge_and_format_cells(dest_sheet, 'K12', 'L12', 'Bucket Truck', alignment="left", font_size=10, isBorder=True)
    merge_and_format_cells(dest_sheet, 'K13', 'L13', 'Bucket Truck 4x4', alignment="left", font_size=10, isBorder=True)
    merge_and_format_cells(dest_sheet, 'K14', 'L14', 'Line Truck', alignment="left", font_size=10, isBorder=True)
    merge_and_format_cells(dest_sheet, 'K15', 'L15', 'Line Truck 4x4', alignment="left", font_size=10, isBorder=True)
    merge_and_format_cells(dest_sheet, 'K16', 'L16', 'Pick-up', alignment="left", font_size=10, isBorder=True)
    merge_and_format_cells(dest_sheet, 'K17', 'L17', 'Digger Derrick', alignment="left", font_size=10, isBorder=True)
    merge_and_format_cells(dest_sheet, 'K18', 'L18', 'Other', alignment="left", font_size=10, isBorder=True)
def getFromPut(dest_sheet, get, put1, put2):
    data = dest_sheet[get].value
    merge_and_format_cells(dest_sheet, put1, put2, data,  alignment='center',  font_size=9, isBorder=True)
    

def copy_and_paste_data():
    source_file_path = "input.xlsm"
    dest_file_path = "output.xlsx"
    source_workbook = openpyxl.load_workbook(source_file_path, keep_vba=True)
    source_sheet = source_workbook.active

    try:
        dest_workbook = openpyxl.load_workbook(dest_file_path, keep_vba=True)
    except FileNotFoundError:
        dest_workbook = Workbook()
        dest_workbook.save(dest_file_path)
        # dest_workbook = load_workbook(dest_file_path, keep_vba=True)

    dest_sheet = dest_workbook.active

    merge_and_format_cells(dest_sheet, 'A1', 'M3', 'PPL Contractor Crewing Worksheet', blueColor, whiteColor, alignment="left", family="head", font_size=28, bold=True)
    subTitle(dest_sheet)
    line(dest_sheet)

    primaryContact(dest_sheet)
    merge_and_format_cells(dest_sheet, 'A10', 'M10', '')


    tableHeading(dest_sheet)
    otherData(dest_sheet)
    
    dest_row = 25
    gf_row = 7

    emp_count = 0
    pich_up_count = 0
    bucket_truck_count = 0
    digger_derrick_count = 0
    other_count = 0

    final_crew = ""

    # Loop through the source column (column C)
    for source_row in range(10, source_sheet.max_row + 1):
        data_to_copy = source_sheet.cell(row=source_row, column=3).value

        if data_to_copy and data_to_copy.strip():
            name_parts = data_to_copy.split()
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""

            dest_sheet.cell(row=dest_row, column=3, value=first_name)
            dest_sheet.cell(row=dest_row, column=2, value=last_name)
            emp_count += 1

            # Copy and paste other columns
            gender = source_sheet.cell(row=source_row, column=5).value
            dest_sheet.cell(row=dest_row, column=12, value=gender)

            phone = source_sheet.cell(row=source_row, column=4).value
            dest_sheet.cell(row=dest_row, column=7, value=phone)

            truck_no = source_sheet.cell(row=source_row, column=12).value
            dest_sheet.cell(row=dest_row, column=11, value=truck_no)

            classification = source_sheet.cell(row=source_row, column=7).value

            if classification in ["General Foreman", "GENERAL FOREMAN"]:
                dest_sheet.cell(row=gf_row, column=3, value=data_to_copy)
                dest_sheet.cell(row=gf_row, column=6, value="GF")
                dest_sheet.cell(row=gf_row, column=7, value=phone)
                gf_row += 1

            classification_map = {
                "Journeyman Lineman": "Line JL",
                "JOURNEYMAN LINEMAN": "Line JL",
                "Operator": "Equipment Op",
                "OPERATOR": "Equipment Op",
                "FOREMAN": "Foreman",
                "GENERAL FOREMAN": "General Foreman",
                "Groundman": "Line Groundhand",
                "DRIVER GROUNDMAN": "Line Groundhand"
            }

            dest_sheet.cell(row=dest_row, column=8, value=classification_map.get(classification, classification))

            crew_num = source_sheet.cell(row=source_row, column=8).value
            final_crew = crew_num if crew_num else final_crew
            dest_sheet.cell(row=dest_row, column=4, value=final_crew)

            state = source_sheet.cell(row=source_row, column=9).value
            dest_sheet.cell(row=dest_row, column=13, value=state)

            dest_sheet.cell(row=dest_row, column=1, value="Premium Utility Contractor")

            equipment = source_sheet.cell(row=source_row, column=11).value
            equipment_map = {
                "3/4 Ton Pickup/1 Ton Pickup": "Pick-up",
                "3/4 TON PICKUP/ 1 TON PICKUP": "Pick-up",
                "55' Bucket Truck": "Bucket Truck",
                "55' BUCKET TRUCK": "Bucket Truck",
                "DIGGER DERRICK": "Digger Derrick",
                "Digger Derrick": "Digger Derrick",
                "Service Truck": "Other"
            }

            equipment_type = equipment_map.get(equipment, equipment)
            dest_sheet.cell(row=dest_row, column=10, value=equipment_type)

            if equipment_type == "Pick-up":
                pich_up_count += 1
            elif equipment_type == "Bucket Truck":
                bucket_truck_count += 1
            elif equipment_type == "Digger Derrick":
                digger_derrick_count += 1
            elif equipment_type == "Other":
                other_count += 1

            dest_row += 1

    dest_sheet.cell(row=12, column=9, value=emp_count)
    dest_sheet.cell(row=13, column=9, value=final_crew)
    dest_sheet.cell(row=12, column=13, value=bucket_truck_count)

    for row in range(25, 201):
        fill_color = lightBlue if row % 2 != 0 else blueBro
        for col in range(1, 24):
            col_letter = get_column_letter(col)
            cell = dest_sheet[f'{col_letter}{row}']
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(color=blackColor, name="Calibri")
    
    for row in range(12, 19):
        dest_sheet[f'M{row}'].border = border
        dest_sheet[f'I{row}'].border = border

    getFromPut(dest_sheet, 'C7', 'C7', 'E7')
    getFromPut(dest_sheet, 'C8', 'C8', 'E8')
    getFromPut(dest_sheet, 'C9', 'C9', 'E9')

    getFromPut(dest_sheet, 'G7', 'G7', 'I7')
    getFromPut(dest_sheet, 'G8', 'G8', 'I8')
    getFromPut(dest_sheet, 'G9', 'G9', 'I9')

    getFromPut(dest_sheet, 'F7', 'F7', 'F7')
    getFromPut(dest_sheet, 'F8', 'F8', 'F8')
    getFromPut(dest_sheet, 'F9', 'F9', 'F9')

    getFromPut(dest_sheet, 'J7', 'J7', 'M7')
    getFromPut(dest_sheet, 'J8', 'J8', 'M8')
    getFromPut(dest_sheet, 'J9', 'J9', 'M9')

    
    # for row in dest_sheet.iter_rows():
    #     for cell in row:
    #         cell.font = default_font
    dest_workbook.save(dest_file_path)
    print(f"Data copied successfully to {dest_file_path}")



# Run the function
copy_and_paste_data()
